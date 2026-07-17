import os
import json
import secrets
import psycopg2
import psycopg2.extras
from functools import wraps
from io import BytesIO
from urllib.parse import urlparse
from flask import Flask, render_template, request, redirect, url_for, jsonify, send_file, session
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta, timezone
from werkzeug.security import generate_password_hash, check_password_hash

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.json.ensure_ascii = False
DATABASE_URL = os.environ.get("DATABASE_URL", "")
EXCEL_FILE = "flota_vehicular.xlsx"


ECUADOR_TZ = timezone(timedelta(hours=-5))

def fecha_local():
    return datetime.now(ECUADOR_TZ).strftime("%Y-%m-%d %H:%M:%S")


@app.template_filter("tojson_filter")
def tojson_filter(obj):
    return json.dumps(dict(obj), ensure_ascii=False, default=str)


def get_db():
    url = urlparse(DATABASE_URL)
    conn = psycopg2.connect(
        database=url.path[1:],
        user=url.username,
        password=url.password,
        host=url.hostname,
        port=url.port,
        sslmode='require'
    )
    conn.autocommit = False
    cur = conn.cursor()
    cur.execute("SET timezone = 'America/Guayaquil'")
    cur.close()
    return conn


def dict_row(cursor):
    colnames = [d[0] for d in cursor.description]
    row = cursor.fetchone()
    if row:
        return dict(zip(colnames, row))
    return None


def dict_rows(cursor):
    colnames = [d[0] for d in cursor.description]
    rows = cursor.fetchall()
    return [dict(zip(colnames, r)) for r in rows]


def init_db():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS usuarios (
            id SERIAL PRIMARY KEY,
            usuario TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            nombre TEXT,
            rol TEXT NOT NULL DEFAULT 'operador',
            fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS vehiculos (
            id SERIAL PRIMARY KEY,
            placa TEXT NOT NULL UNIQUE,
            chasis TEXT,
            motor TEXT,
            marca TEXT,
            modelo TEXT,
            anio INTEGER,
            kilometraje INTEGER DEFAULT 0,
            ubicacion TEXT,
            estado TEXT DEFAULT 'Activo',
            creado_por TEXT,
            mecanica TEXT DEFAULT 'Multimarcas',
            fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS mantenimientos (
            id SERIAL PRIMARY KEY,
            vehiculo_id INTEGER NOT NULL,
            fecha TEXT NOT NULL,
            tipo TEXT,
            descripcion TEXT,
            kilometraje INTEGER,
            costo REAL DEFAULT 0,
            taller TEXT,
            FOREIGN KEY (vehiculo_id) REFERENCES vehiculos(id) ON DELETE CASCADE
        )
    """)
    cur.execute("""
        CREATE TABLE IF NOT EXISTS parte_taller (
            id SERIAL PRIMARY KEY,
            vehiculo_id INTEGER NOT NULL,
            placa TEXT NOT NULL,
            mecanica TEXT,
            motivo TEXT,
            observaciones TEXT,
            fecha_ingreso TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            fecha_salida TIMESTAMP,
            registrado_por TEXT,
            FOREIGN KEY (vehiculo_id) REFERENCES vehiculos(id) ON DELETE CASCADE
        )
    """)
    cur.execute("SELECT id FROM usuarios WHERE rol='admin'")
    admin = cur.fetchone()
    if not admin:
        cur.execute(
            "INSERT INTO usuarios (usuario, password, nombre, rol) VALUES (%s, %s, %s, %s)",
            ("admin", generate_password_hash("Willian098"), "Administrador", "admin")
        )
    conn.commit()
    cur.close()
    conn.close()


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login"))
        if session.get("rol") != "admin":
            return redirect(url_for("index"))
        return f(*args, **kwargs)
    return decorated


@app.route("/login", methods=["GET", "POST"])
def login():
    error = None
    if request.method == "POST":
        usuario = request.form["usuario"].strip()
        password = request.form["password"]
        conn = get_db()
        cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute("SELECT * FROM usuarios WHERE usuario=%s", (usuario,))
        user = cur.fetchone()
        cur.close()
        conn.close()
        if user and check_password_hash(user["password"], password):
            session["user_id"] = user["id"]
            session["usuario"] = user["usuario"]
            session["nombre"] = user["nombre"]
            session["rol"] = user["rol"]
            return redirect(url_for("index"))
        error = "Usuario o contrasena incorrectos"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/usuarios")
@admin_required
def gestionar_usuarios():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM usuarios ORDER BY id")
    usuarios = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("usuarios.html", usuarios=usuarios)


@app.route("/usuarios/crear", methods=["POST"])
@admin_required
def crear_usuario():
    usuario = request.form["usuario"].strip()
    password = request.form["password"]
    nombre = request.form["nombre"].strip()
    rol = request.form["rol"]
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute(
            "INSERT INTO usuarios (usuario, password, nombre, rol) VALUES (%s, %s, %s, %s)",
            (usuario, generate_password_hash(password), nombre, rol)
        )
        conn.commit()
    except psycopg2.IntegrityError:
        conn.rollback()
    cur.close()
    conn.close()
    return redirect(url_for("gestionar_usuarios"))


@app.route("/usuarios/eliminar/<int:id>")
@admin_required
def eliminar_usuario(id):
    if id == session.get("user_id"):
        return redirect(url_for("gestionar_usuarios"))
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM usuarios WHERE id=%s AND rol != 'admin'", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("gestionar_usuarios"))


@app.route("/")
@login_required
def index():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    mecanica_filter = request.args.get("mecanica", "")
    if mecanica_filter:
        cur.execute("SELECT * FROM vehiculos WHERE mecanica=%s ORDER BY id DESC", (mecanica_filter,))
    else:
        cur.execute("SELECT * FROM vehiculos ORDER BY id DESC")
    vehiculos = cur.fetchall()
    cur.execute("SELECT DISTINCT mecanica FROM vehiculos WHERE mecanica IS NOT NULL AND mecanica != '' ORDER BY mecanica")
    mecanicas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("index.html", vehiculos=vehiculos, mecanicas=mecanicas, mecanica_actual=mecanica_filter)


@app.route("/agregar", methods=["POST"])
@login_required
def agregar():
    conn = get_db()
    cur = conn.cursor()
    try:
        cur.execute("""
            INSERT INTO vehiculos (placa, chasis, motor, marca, modelo, anio, kilometraje, ubicacion, estado, creado_por, mecanica, fecha_registro)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (
            request.form["placa"].upper(),
            request.form["chasis"],
            request.form["motor"],
            request.form["marca"],
            request.form["modelo"],
            request.form["anio"],
            request.form["kilometraje"],
            request.form["ubicacion"],
            request.form["estado"],
            session.get("usuario", ""),
            request.form.get("mecanica", "Multimarcas"),
            fecha_local()
        ))
        conn.commit()
    except psycopg2.IntegrityError:
        conn.rollback()
    finally:
        cur.close()
        conn.close()
    return redirect(url_for("index"))


@app.route("/editar/<int:id>", methods=["POST"])
@admin_required
def editar(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        UPDATE vehiculos
        SET placa=%s, chasis=%s, motor=%s, marca=%s, modelo=%s, anio=%s, kilometraje=%s, ubicacion=%s, estado=%s, mecanica=%s
        WHERE id=%s
    """, (
        request.form["placa"].upper(),
        request.form["chasis"],
        request.form["motor"],
        request.form["marca"],
        request.form["modelo"],
        request.form["anio"],
        request.form["kilometraje"],
        request.form["ubicacion"],
        request.form["estado"],
        request.form.get("mecanica", "Multimarcas"),
        id
    ))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("index"))


@app.route("/eliminar/<int:id>")
@admin_required
def eliminar(id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM vehiculos WHERE id=%s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("index"))


@app.route("/vehiculo/<int:id>")
@login_required
def detalle(id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM vehiculos WHERE id=%s", (id,))
    vehiculo = cur.fetchone()
    cur.execute("SELECT * FROM mantenimientos WHERE vehiculo_id=%s ORDER BY fecha DESC", (id,))
    historial = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("detalle.html", vehiculo=vehiculo, historial=historial)


@app.route("/mantenimiento/<int:vehiculo_id>", methods=["POST"])
@login_required
def agregar_mantenimiento(vehiculo_id):
    conn = get_db()
    cur = conn.cursor()
    cur.execute("""
        INSERT INTO mantenimientos (vehiculo_id, fecha, tipo, descripcion, kilometraje, costo, taller)
        VALUES (%s, %s, %s, %s, %s, %s, %s)
    """, (
        vehiculo_id,
        request.form["fecha"],
        request.form["tipo"],
        request.form["descripcion"],
        request.form["kilometraje"],
        request.form["costo"],
        request.form["taller"]
    ))
    cur.execute("UPDATE vehiculos SET kilometraje=%s WHERE id=%s", (request.form["kilometraje"], vehiculo_id))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("detalle", id=vehiculo_id))


@app.route("/mantenimiento/editar/<int:id>", methods=["POST"])
@admin_required
def editar_mantenimiento(id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT vehiculo_id FROM mantenimientos WHERE id=%s", (id,))
    mant = cur.fetchone()
    if not mant:
        cur.close()
        conn.close()
        return redirect(url_for("index"))
    vehiculo_id = mant["vehiculo_id"]
    cur.execute("""
        UPDATE mantenimientos
        SET fecha=%s, tipo=%s, descripcion=%s, kilometraje=%s, costo=%s, taller=%s
        WHERE id=%s
    """, (
        request.form["fecha"],
        request.form["tipo"],
        request.form["descripcion"],
        request.form["kilometraje"],
        request.form["costo"],
        request.form["taller"],
        id
    ))
    cur.execute("UPDATE vehiculos SET kilometraje=%s WHERE id=%s", (request.form["kilometraje"], vehiculo_id))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("detalle", id=vehiculo_id))


@app.route("/mantenimiento/eliminar/<int:id>")
@admin_required
def eliminar_mantenimiento(id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT vehiculo_id FROM mantenimientos WHERE id=%s", (id,))
    mant = cur.fetchone()
    if not mant:
        cur.close()
        conn.close()
        return redirect(url_for("index"))
    vehiculo_id = mant["vehiculo_id"]
    cur.execute("DELETE FROM mantenimientos WHERE id=%s", (id,))
    conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("detalle", id=vehiculo_id))


@app.route("/buscar")
@login_required
def buscar():
    q = request.args.get("q", "")
    mecanica_filter = request.args.get("mecanica", "")
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    query = "SELECT * FROM vehiculos WHERE (placa ILIKE %s OR marca ILIKE %s OR ubicacion ILIKE %s OR chasis ILIKE %s)"
    params = [f"%{q}%", f"%{q}%", f"%{q}%", f"%{q}%"]
    if mecanica_filter:
        query += " AND mecanica=%s"
        params.append(mecanica_filter)
    query += " ORDER BY id DESC"
    cur.execute(query, params)
    vehiculos = cur.fetchall()
    cur.execute("SELECT DISTINCT mecanica FROM vehiculos WHERE mecanica IS NOT NULL AND mecanica != '' ORDER BY mecanica")
    mecanicas = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("index.html", vehiculos=vehiculos, busqueda=q, mecanicas=mecanicas, mecanica_actual=mecanica_filter)


@app.route("/api/estadisticas")
@login_required
def estadisticas():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("SELECT COUNT(*) FROM vehiculos")
    total = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vehiculos WHERE estado='Activo'")
    activos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vehiculos WHERE estado='Inactivo'")
    inactivos = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vehiculos WHERE estado='En Taller'")
    taller = cur.fetchone()[0]
    cur.execute("SELECT COUNT(*) FROM vehiculos WHERE estado='Remate'")
    remate = cur.fetchone()[0]
    cur.close()
    conn.close()
    return jsonify({
        "total": total,
        "activos": activos,
        "inactivos": inactivos,
        "en_taller": taller,
        "remate": remate
    })


@app.route("/exportar-excel")
@login_required
def exportar():
    buffer = exportar_excel()
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"flota_vehicular_{timestamp}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/exportar-vehiculo/<int:id>")
@login_required
def exportar_vehiculo(id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT * FROM vehiculos WHERE id=%s", (id,))
    vehiculo = cur.fetchone()
    if not vehiculo:
        cur.close()
        conn.close()
        return redirect(url_for("index"))
    cur.execute("SELECT * FROM mantenimientos WHERE vehiculo_id=%s ORDER BY fecha DESC", (id,))
    historial = cur.fetchall()
    cur.close()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = vehiculo["placa"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    fill_blue = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    fill_green = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")

    info = [
        ("Placa", vehiculo["placa"]), ("Marca", vehiculo["marca"]),
        ("Modelo", vehiculo["modelo"]), ("Anio", vehiculo["anio"]),
        ("Chasis", vehiculo["chasis"]), ("Motor", vehiculo["motor"]),
        ("Kilometraje", vehiculo["kilometraje"]), ("Ubicacion", vehiculo["ubicacion"]),
        ("Estado", vehiculo["estado"]), ("Mecanica", vehiculo["mecanica"])
    ]
    for row_idx, (label, val) in enumerate(info, 1):
        cell_label = ws.cell(row=row_idx, column=1, value=label)
        cell_label.font = Font(bold=True, color="FFFFFF")
        cell_label.fill = fill_blue
        cell_label.border = thin_border
        cell_val = ws.cell(row=row_idx, column=2, value=val or "-")
        cell_val.border = thin_border
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 30

    mant_row = len(info) + 3
    ws.cell(row=mant_row, column=1, value="HISTORIAL DE MANTENIMIENTO").font = Font(bold=True, size=14, color="22C55E")
    ws.merge_cells(start_row=mant_row, start_column=1, end_row=mant_row, end_column=7)
    mant_row += 1

    mant_headers = ["Fecha", "Tipo", "Descripcion", "Kilometraje", "Costo", "Taller"]
    for col, h in enumerate(mant_headers, 1):
        cell = ws.cell(row=mant_row, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_green
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    mant_row += 1

    for m in historial:
        for col, key in enumerate(["fecha", "tipo", "descripcion", "kilometraje", "costo", "taller"], 1):
            val = m[key] if m[key] else "-"
            cell = ws.cell(row=mant_row, column=col, value=val)
            cell.border = thin_border
            if key == "costo":
                cell.number_format = "#,##0.00"
            if key == "kilometraje":
                cell.number_format = "#,##0"
        mant_row += 1

    if not historial:
        ws.cell(row=mant_row, column=1, value="Sin registros de mantenimiento").font = Font(italic=True, color="64748B")

    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        as_attachment=True,
        download_name=f"{vehiculo['placa']}_historial.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/importar-excel", methods=["POST"])
@admin_required
def importar():
    if "archivo" not in request.files:
        return redirect(url_for("index"))
    file = request.files["archivo"]
    if file.filename == "":
        return redirect(url_for("index"))
    if not file.filename.endswith((".xlsx", ".xls")):
        return redirect(url_for("index"))
    filepath = os.path.join("uploads", file.filename)
    os.makedirs("uploads", exist_ok=True)
    file.save(filepath)
    imported, updated = importar_excel(filepath)
    os.remove(filepath)
    return redirect(url_for("index"))


@app.route("/sync-excel", methods=["POST"])
@login_required
def sync_excel():
    return jsonify({"status": "ok", "message": "Sync no disponible en version online"})


def exportar_excel():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    wb = Workbook()
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    veh_headers = ["ID", "Placa", "Chasis", "Motor", "Marca", "Modelo", "Anio", "Kilometraje", "Ubicacion", "Estado", "Mecanica", "Creado Por", "Fecha Registro"]
    veh_cols = ["id", "placa", "chasis", "motor", "marca", "modelo", "anio", "kilometraje", "ubicacion", "estado", "mecanica", "creado_por", "fecha_registro"]
    mant_headers = ["ID", "Vehiculo ID", "Placa", "Fecha", "Tipo", "Descripcion", "Kilometraje", "Costo", "Taller"]
    mant_cols = ["id", "vehiculo_id", "placa", "fecha", "tipo", "descripcion", "kilometraje", "costo", "taller"]
    fill_mant = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")

    TODAS_LAS_MECANICAS = [
        "Multimarcas", "Kia", "Ambacar", "Chevrolet", "Pesados",
        "Great Wall 2025", "Great Wall 2026", "Motos Multimarcas",
        "Honda", "Morini", "Kia Tasman"
    ]

    cur.execute("SELECT * FROM vehiculos ORDER BY id")
    vehiculos = cur.fetchall()
    ws_all = wb.active
    ws_all.title = "Todos los Vehiculos"
    fill_all = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    for col, h in enumerate(veh_headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_all
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    for row_idx, v in enumerate(vehiculos, 2):
        for col_idx, key in enumerate(veh_cols, 1):
            val = v.get(key, "") if v else ""
            cell = ws_all.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if key == "kilometraje":
                cell.number_format = "#,##0"
    for col in range(1, len(veh_headers) + 1):
        ws_all.column_dimensions[get_column_letter(col)].width = 18

    colores = ["16A34A", "EA580C", "7C3AED", "DC2626", "0891B2", "CA8A04", "DB2777", "4F46E5", "059669", "9333EA", "B91C1C"]

    for idx, nombre in enumerate(TODAS_LAS_MECANICAS):
        ws = wb.create_sheet(title=nombre[:31])
        fill = PatternFill(start_color=colores[idx], end_color=colores[idx], fill_type="solid")

        cur.execute("SELECT * FROM vehiculos WHERE mecanica=%s ORDER BY id", (nombre,))
        vehiculos_mec = cur.fetchall()

        ws.cell(row=1, column=1, value=f"VEHICULOS - {nombre.upper()}").font = Font(bold=True, size=14, color=colores[idx])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(veh_headers))

        for col, h in enumerate(veh_headers, 1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        row_start = 3
        for row_idx, v in enumerate(vehiculos_mec):
            for col_idx, key in enumerate(veh_cols, 1):
                val = v.get(key, "") if v else ""
                cell = ws.cell(row=row_start + row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if key == "kilometraje":
                    cell.number_format = "#,##0"

        if not vehiculos_mec:
            ws.cell(row=3, column=1, value="Sin vehiculos registrados").font = Font(italic=True, color="64748B")

        mant_row = row_start + len(vehiculos_mec) + 2
        cur.execute("""
            SELECT m.*, v.placa FROM mantenimientos m
            LEFT JOIN vehiculos v ON m.vehiculo_id = v.id
            WHERE v.mecanica=%s
            ORDER BY m.fecha DESC
        """, (nombre,))
        mant_mec = cur.fetchall()

        ws.cell(row=mant_row, column=1, value=f"HISTORIAL MANTENIMIENTO - {nombre.upper()}").font = Font(bold=True, size=14, color="22C55E")
        ws.merge_cells(start_row=mant_row, start_column=1, end_row=mant_row, end_column=len(mant_headers))
        mant_row += 1

        for col, h in enumerate(mant_headers, 1):
            cell = ws.cell(row=mant_row, column=col, value=h)
            cell.font = header_font
            cell.fill = fill_mant
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        mant_row += 1

        for row_idx, m in enumerate(mant_mec):
            for col_idx, key in enumerate(mant_cols, 1):
                val = m.get(key, "") if m else ""
                cell = ws.cell(row=mant_row + row_idx, column=col_idx, value=val)
                cell.border = thin_border
                if key == "costo":
                    cell.number_format = "#,##0.00"

        if not mant_mec:
            ws.cell(row=mant_row, column=1, value="Sin mantenimientos registrados").font = Font(italic=True, color="64748B")

        for col in range(1, max(len(veh_headers), len(mant_headers)) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18

    ws_mant_all = wb.create_sheet("Todos Mantenimientos")
    for col, h in enumerate(mant_headers, 1):
        cell = ws_mant_all.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = fill_mant
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border
    cur.execute("""
        SELECT m.*, v.placa FROM mantenimientos m
        LEFT JOIN vehiculos v ON m.vehiculo_id = v.id
        ORDER BY m.fecha DESC
    """)
    mant_all = cur.fetchall()
    for row_idx, m in enumerate(mant_all, 2):
        for col_idx, key in enumerate(mant_cols, 1):
            val = m.get(key, "") if m else ""
            cell = ws_mant_all.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border
            if key == "costo":
                cell.number_format = "#,##0.00"
    for col in range(1, len(mant_headers) + 1):
        ws_mant_all.column_dimensions[get_column_letter(col)].width = 20

    cur.close()
    conn.close()
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def importar_excel(filepath):
    wb = load_workbook(filepath)
    ws = wb["Vehiculos"]
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    imported = 0
    updated = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        placa = str(row[1]).upper().strip()
        cur.execute("SELECT id FROM vehiculos WHERE placa=%s", (placa,))
        existing = cur.fetchone()
        data = {
            "chasis": row[2] or "", "motor": row[3] or "", "marca": row[4] or "",
            "modelo": row[5] or "", "anio": row[6] or 0, "kilometraje": row[7] or 0,
            "ubicacion": row[8] or "", "estado": row[9] or "Activo", "mecanica": row[10] or "Multimarcas"
        }
        if existing:
            cur.execute("""
                UPDATE vehiculos SET chasis=%s, motor=%s, marca=%s, modelo=%s, anio=%s, kilometraje=%s, ubicacion=%s, estado=%s, mecanica=%s
                WHERE placa=%s
            """, (data["chasis"], data["motor"], data["marca"], data["modelo"], data["anio"], data["kilometraje"], data["ubicacion"], data["estado"], data["mecanica"], placa))
            updated += 1
        else:
            cur.execute("""
                INSERT INTO vehiculos (placa, chasis, motor, marca, modelo, anio, kilometraje, ubicacion, estado, mecanica)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (placa, data["chasis"], data["motor"], data["marca"], data["modelo"], data["anio"], data["kilometraje"], data["ubicacion"], data["estado"], data["mecanica"]))
            imported += 1
    if "Mantenimientos" in wb.sheetnames:
        ws2 = wb["Mantenimientos"]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row[1]:
                continue
            vehiculo_id = row[1]
            cur.execute("SELECT id FROM vehiculos WHERE id=%s", (vehiculo_id,))
            v = cur.fetchone()
            if v:
                cur.execute("""
                    INSERT INTO mantenimientos (vehiculo_id, fecha, tipo, descripcion, kilometraje, costo, taller)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                """, (vehiculo_id, str(row[3] or ""), row[4] or "", row[5] or "", row[6] or 0, row[7] or 0, row[8] or ""))
    conn.commit()
    cur.close()
    conn.close()
    return imported, updated


@app.route("/parte-taller")
@login_required
def parte_taller():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    filtro = request.args.get("filtro", "activos")
    if filtro == "salidos":
        cur.execute("""
            SELECT pt.*, v.marca, v.modelo
            FROM parte_taller pt
            LEFT JOIN vehiculos v ON pt.vehiculo_id = v.id
            WHERE pt.fecha_salida IS NOT NULL
            ORDER BY pt.fecha_salida DESC LIMIT 100
        """)
    else:
        cur.execute("""
            SELECT pt.*, v.marca, v.modelo
            FROM parte_taller pt
            LEFT JOIN vehiculos v ON pt.vehiculo_id = v.id
            WHERE pt.fecha_salida IS NULL
            ORDER BY pt.fecha_ingreso DESC
        """)
    registros = cur.fetchall()
    cur.execute("SELECT id, placa, marca, modelo FROM vehiculos ORDER BY placa")
    vehiculos = cur.fetchall()
    cur.close()
    conn.close()
    return render_template("parte_taller.html", registros=registros, vehiculos=vehiculos, filtro=filtro)


@app.route("/parte-taller/ingresar", methods=["POST"])
@login_required
def parte_taller_ingresar():
    conn = get_db()
    cur = conn.cursor()
    vehiculo_id = request.form["vehiculo_id"]
    cur.execute("SELECT placa, mecanica FROM vehiculos WHERE id=%s", (vehiculo_id,))
    v = cur.fetchone()
    if v:
        cur.execute("""
            INSERT INTO parte_taller (vehiculo_id, placa, mecanica, motivo, observaciones, fecha_ingreso, registrado_por)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        """, (vehiculo_id, v[0], v[1], request.form["motivo"], request.form.get("observaciones", ""), fecha_local(), session.get("usuario", "")))
        cur.execute("UPDATE vehiculos SET estado='En Taller' WHERE id=%s", (vehiculo_id,))
        conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("parte_taller"))


@app.route("/parte-taller/salir/<int:id>")
@login_required
def parte_taller_salir(id):
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
    cur.execute("SELECT vehiculo_id FROM parte_taller WHERE id=%s AND fecha_salida IS NULL", (id,))
    reg = cur.fetchone()
    if reg:
        cur.execute("UPDATE parte_taller SET fecha_salida=%s WHERE id=%s", (fecha_local(), id))
        cur.execute("UPDATE vehiculos SET estado='Activo' WHERE id=%s", (reg["vehiculo_id"],))
        conn.commit()
    cur.close()
    conn.close()
    return redirect(url_for("parte_taller"))


@app.route("/reportes")
@login_required
def reportes():
    conn = get_db()
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    mecanica_filter = request.args.get("mecanica", "")
    tipo_filter = request.args.get("tipo", "")
    fecha_desde = request.args.get("fecha_desde", "")
    fecha_hasta = request.args.get("fecha_hasta", "")

    cur.execute("SELECT DISTINCT mecanica FROM vehiculos WHERE mecanica IS NOT NULL AND mecanica != '' ORDER BY mecanica")
    mecanicas = cur.fetchall()

    query = """
        SELECT m.*, v.placa, v.mecanica, v.marca, v.modelo
        FROM mantenimientos m
        LEFT JOIN vehiculos v ON m.vehiculo_id = v.id
        WHERE 1=1
    """
    params = []

    if mecanica_filter:
        query += " AND v.mecanica = %s"
        params.append(mecanica_filter)
    if tipo_filter:
        query += " AND m.tipo = %s"
        params.append(tipo_filter)
    if fecha_desde:
        query += " AND m.fecha >= %s"
        params.append(fecha_desde)
    if fecha_hasta:
        query += " AND m.fecha <= %s"
        params.append(fecha_hasta)

    query += " ORDER BY m.fecha DESC"
    cur.execute(query, params)
    mantenimientos = cur.fetchall()

    cur.execute("SELECT DISTINCT tipo FROM mantenimientos WHERE tipo IS NOT NULL AND tipo != '' ORDER BY tipo")
    tipos = cur.fetchall()

    total = len(mantenimientos)
    correctivos = sum(1 for m in mantenimientos if m.get("tipo") == "Correctivo")
    preventivos = sum(1 for m in mantenimientos if m.get("tipo") == "Preventivo")
    otro = total - correctivos - preventivos
    costo_total = sum(m.get("costo") or 0 for m in mantenimientos)

    por_tipo = {}
    for m in mantenimientos:
        t = m.get("tipo") or "Sin tipo"
        por_tipo[t] = por_tipo.get(t, 0) + 1

    por_mecanica = {}
    for m in mantenimientos:
        mc = m.get("mecanica") or "Sin mecanica"
        por_mecanica[mc] = por_mecanica.get(mc, 0) + 1

    cur.close()
    conn.close()
    return render_template("reportes.html",
        mantenimientos=mantenimientos,
        mecanicas=mecanicas,
        tipos=tipos,
        total=total,
        correctivos=correctivos,
        preventivos=preventivos,
        otro=otro,
        costo_total=costo_total,
        por_tipo=por_tipo,
        por_mecanica=por_mecanica,
        mecanica_actual=mecanica_filter,
        tipo_actual=tipo_filter,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta
    )


init_db()


@app.route("/health")
def health():
    return jsonify({"status": "ok"})


if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
